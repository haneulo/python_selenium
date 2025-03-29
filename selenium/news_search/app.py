# 네이버 뉴스 크롤링 후 엑셀저장 프로그램
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
import chromedriver_autoinstaller
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

chromedriver_autoinstaller.install()

driver = webdriver.Chrome()

search_query = "속보"
search_link = f"https://search.naver.com/search.naver?where=news&ie=utf8&sm=nws_hty&query={search_query}"

driver.get(search_link)

time.sleep(1)

# 최신순으로 보기
Latest_selector = "#snb > div.mod_group_option_filter._search_option_simple_wrap > div > div.option_area.type_sort > a:nth-child(2)"
Latest_element = driver.find_element(By.CSS_SELECTOR, Latest_selector)
Latest_element.click()

time.sleep(1)

# 스크롤 동작(스크롤 1번당 10개의 뉴스 추가)
for _ in range(4):
    driver.execute_script("window.scrollBy(0,10000);")
    time.sleep(1)


req = driver.page_source
soup = BeautifulSoup(req, 'html.parser')

# 뉴스 타이틀, url, 언론사 정보 select
articles = soup.select('.news_wrap')


# 엑셀 파일 생성
wb = Workbook()
ws = wb.active
ws.title = "news_list"
ws.append(["제목", "링크", "언론사"])

# 검색결과를 엑셀에 넣기
for article in articles:
    title = article.select_one('.news_tit').text
    url = article.select_one('.news_tit')['href']
    company = article.select_one('.info_group > a').text.split(' ')[0].replace('언론사','')

    ws.append([title, url, company])
    
# 엑셀 열 가로길이 변경
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 15

# 셀 가운데정렬, bold, 배경색 설정
for row_range in ws['A1':'C1']:
    for cell in row_range:
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='fff5cc', fill_type = 'solid')

# 엑셀 저장
wb.save("news.xlsx")
print("검색결과 저장이 완료되었습니다.")

driver.quit()
